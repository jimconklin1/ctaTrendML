import pandas as pd
import numpy as np
import argparse
import openpyxl.styles as opxs
import operator
import os
import sys
import dateutil
import warnings
import shutil
import dateutil.relativedelta
import openpyxl as pxl
from openpyxl.utils.dataframe import dataframe_to_rows

from util.eh_ml_io import load_ml
from util.eh_types import *
from util.basic_util import *
from util.eh_util import *
from util.eh_db_util import *
from util.redemption import *
from util.opx_util import *
from util.sig import *
from util.const import *
from eh.inv_obj import *

ignore_missing_returns = True

ml_data_path = "D:/Local/PubEq/MLData/EH2"
ml_fn_suff = "1mo"
output_path = "{}/{}/{}".format(ml_data_path, "out", ml_fn_suff) 
if not os.path.exists(output_path): os.mkdir(output_path)

ml_import_dir = "{}/ml_to_python".format(ml_data_path)
ml = load_ml("{}/eh2_{}.mat".format(ml_import_dir, ml_fn_suff))
if ml.months_per_period not in [1,3]:
    raise ValueError("{}-month period length is not supported".format(ml.months_per_period))

ref = argparse.Namespace()
ref_pickle_fn = "{}/ref.pkl".format(ml_import_dir)
if os.path.isfile(ref_pickle_fn):
    ref.df = pd.read_pickle(ref_pickle_fn)
else:    
    db = connect_to_cfg_db("PubEqCoreProd_RO")
    ref.df = pd.read_sql("select * from eh.fund", db, parse_dates=['INCEPTION_DATE'], index_col='FUND_ID')
    ref.df.to_pickle(ref_pickle_fn)

ref.fund_ids = ml.db_data.fund_ids
ref.dates = ml.calc.dates
ref.pos_dates = ref.dates.append(pd.DatetimeIndex([ref.dates[-1] +dateutil.relativedelta.relativedelta(months=1)]))

cfg = get_config()
assert cfg.entry.aum_mm >= ml.db_data.aum_ever_achieved_mm, "Matlab data was prepared with AUM = ${}mm, which is too large for your filters. You need to rerun Matlab code.".format(round(ml.db_data.aum_ever_achieved_mm))
cfg.inv.months_per_period = ml.months_per_period

ref_len = len(ref.df)
ref.df = parse_redemption_info(ref.df, cfg.inv.exit.delay, cfg.inv.exit.delay)
assert ref_len == len(ref.df), "Lost reference data records!"

ref.df = ref.df.assign(redemp_lag_mo = np.maximum(ref.df['redemp_freq_len_mo'] / 2 + ref.df['redemp_notif_len_mo'], 1))

ref.df = enrich_ref(ref.df.loc[ref.fund_ids,:], ml)
assert len(ref.df) == len(ref.fund_ids), "Some ref records were lost after filtering!"

valid_incept_dt = ~pd.isna(ref.df.INCEPTION_DATE)
cfg.min_invest_dt = ref.df.INCEPTION_DATE.copy()
with warnings.catch_warnings():
    warnings.filterwarnings('ignore', r'Adding/subtracting object-dtype array to DatetimeArray not vectorized')
    cfg.min_invest_dt[valid_incept_dt] += np.tile(
        dateutil.relativedelta.relativedelta(months=cfg.entry.track_rec_mo), np.sum(valid_incept_dt))

sig_flt = build_signals(cfg, ml)
sig=argparse.Namespace()
sig.entry = combine_filter_map(sig_flt.entry_map, operator.and_)
sig.exit = combine_filter_map(sig_flt.exit_map, operator.or_)

ever_enter_trg_flt = sum(sig.entry)>0
ref.fund_flt = ever_enter_trg_flt
if cfg.ccy_flt is not None:
    orig_fund_id_idx = pd.DataFrame(index=ref.fund_ids).index
    filtered_id_idx = ref.df.loc[ref.df.CURRENCY.isin(cfg.ccy_flt)].index
    ref.fund_flt &= orig_fund_id_idx.isin(filtered_id_idx)

compact_sig = filter_attr_cols_2d(sig, ref.fund_flt)

ret_df = resample_returns(
    ml.db_data.equHFrtns.iloc[(ml.db_data.equHFrtns.index >= ml.first_period_dt) & (ml.db_data.equHFrtns.index <= ml.last_period_dt)
    , ref.fund_flt], cfg.inv.months_per_period)

if cfg.inv.exit.is_uniform_delay:
    exit_delay_arr = np.tile(cfg.inv.exit.delay, (1,compact_sig.exit.shape[1]))
else:  
    exit_delay_arr = np.expand_dims(np.ceil(ref.df.loc[ref.fund_flt,'redemp_lag_mo'].to_numpy()), axis=0) + cfg.inv.exit.processing_time

inv = argparse.Namespace(cfg = cfg.inv, sig = compact_sig, exit_delay = exit_delay_arr
    , entry_delay = np.tile(cfg.inv.entry.delay, (1,compact_sig.entry.shape[1])))

portf = FundIndexBuilder(inv, ref)
portf.run()
pos_seq = cumsum_with_reset(portf.pos !=0)
trunc_weights = portf.weights[:-1, :]

missing_ret_flg = pd.isna(ret_df) & (trunc_weights>0)
have_ret_flg = (~pd.isna(ret_df)) & (trunc_weights>0)
miss_ret_str_map = ( pd.DataFrame(bool_mat_to_str_ind(missing_ret_flg, 'x')) + pd.DataFrame(bool_mat_to_str_ind(have_ret_flg, '1')) ).to_numpy()
validate_missing_returns(miss_ret_str_map, ref.fund_ids, ref.fund_flt)
ret_for_calc = ret_df.to_numpy()
# now that we have validated that there are no holes in returns (only trailing or leading blanks),
# we can replace nans with zeroes: this is following our assumption that we can exit right away
# as returns stopped coming
if miss_ret_str_map.size>0:
    ret_for_calc[miss_ret_str_map=='x'] = 0
miss_ret_flt = np.count_nonzero(missing_ret_flg,0)>0
miss_ret_rpt_flt = miss_ret_flt.copy()
if ignore_missing_returns:
    miss_ret_flt = np.tile(False, miss_ret_flt.shape)

if cfg.rpt.style_by_period: inv_stat_qtr_sum = sum_val_by_period(portf.weights, ref, style_fld)

portf_ret = portf_returns(trunc_weights, ret_for_calc, flt=~miss_ret_flt)

flt_str_data = flt_str_expand({k:v.filter[:,ref.fund_flt] for k, v in sig_flt.exit_map.items()})
first_ret_idx = np.where(~np.isnan(portf_ret))[0][0]
portf_ret_df = pd.DataFrame(portf_ret, index=ref.dates, columns = ['Raw_Ret'])
portf_ret_df = portf_ret_df.iloc[first_ret_idx:,:]
portf_ret_df = portf_ret_df.assign(Raw_CmlRet= (portf_ret_df['Raw_Ret'] +1).cumprod()-1)

# -------------------------------
# Prepare output and write to CSVs
# -------------------------------

def enrich_and_write(arr, file_name, **kwargs):
    return enrich_and_write_ex(arr, ref, output_path, file_name, **kwargs)

pos_col_titles=ref.pos_dates.strftime('%Y-%m')
rpt = argparse.Namespace()
rpt.entry_rpt = enrich_and_write(compact_sig.entry, "sig_entry.csv")
rpt.exit_rpt = enrich_and_write(compact_sig.exit, "sig_exit.csv")
rpt.exit_debug_rpt = enrich_and_write(flt_str_data, "sig_exit_STR.csv")
rpt.pos_rpt = enrich_and_write(pos_seq, "pos.csv", val_map={0:np.nan}, columns=pos_col_titles)
rpt.miss_ret_flg_rpt = enrich_and_write(miss_ret_str_map, "miss_ret_flg.csv", flt = miss_ret_rpt_flt, add_flds = ['DEAD'])
rpt.miss_ret_rpt = enrich_and_write(ret_df.to_numpy(), "miss_ret.csv", flt = miss_ret_rpt_flt, add_flds = ['DEAD'])
rpt.wgt_rpt = enrich_and_write(np.where(portf.weights==0, np.nan, portf.weights), "wgt.csv", flds = [style_fld], columns=pos_col_titles)

run_sum_fn = "{}/run_sum.csv".format(output_path)
#if os.path.isfile(run_sum_fn):
#    shutil.move(run_sum_fn, "{}/run_sum_prev.csv".format(output_path))

run_sum_df = prep_run_summary(portf_ret_df, portf.stat, trunc_weights, rpt)
run_sum_highlight = [2] if cfg.output.highlight_mean_return_csv else None
write_df_to_csv(run_sum_df, run_sum_fn, highlight_lines=run_sum_highlight, csv_params = dict(float_format='%.4f', index=False))

tot_periods_for_fund = np.sum(trunc_weights>0, 0)
ref_wgt_df = ref.df.loc[ref.fund_flt].assign(wgt=tot_periods_for_fund / np.sum(tot_periods_for_fund))
style_sum_df = build_fld_summary(ref_wgt_df, style_fld)
style_sum_df.to_csv("{}/sum_style.csv".format(output_path), index=False, float_format='%.4f')
ccy_sum_df = build_fld_summary(ref_wgt_df, 'CURRENCY')
ccy_sum_df.to_csv("{}/sum_ccy.csv".format(output_path), index=False, float_format='%.4f')

portf_ret_df.to_csv("{}/portf_ret.csv".format(output_path), index=True)

model_spec_sum = get_model_spec(sig_flt, cfg, portf.style_wgt_range)
np.savetxt("{}/model_spec.txt".format(output_path), model_spec_sum, '%s')
save_model_spec_code(sys.argv[0], output_path)

if cfg.rpt.style_by_period:
    inv_stat_qtr_sum.cnt_df.to_csv("{}/inv_stat_cnt.csv".format(output_path), float_format='%.4f')
    inv_stat_qtr_sum.frac_df.to_csv("{}/inv_stat_frac.csv".format(output_path), float_format='%.4f')


# -------------------------------
# Consolidated Excel model report
# -------------------------------

if cfg.rpt.excel:
    wb = pxl.Workbook()
    ws_sum = wb.active
    if ws_sum is None: 
        ws_sum = wb.create_sheet(title="Summary")
    else:    
        ws_sum.title = "Summary"
    ws_spec = wb.create_sheet(title="Spec")
    for s in model_spec_sum: ws_spec.append([s])

    date_style = opxs.NamedStyle(name='datetime', number_format='MM/DD/YYYY')

    output_std_rpt(wb, rpt.entry_rpt, 'Entry')
    output_std_rpt(wb, rpt.exit_rpt, 'Exit')
    output_std_rpt(wb, rpt.pos_rpt, 'Pos')
    output_std_rpt(wb, rpt.exit_debug_rpt, 'ExtDbg')

    ws_ret = wb.create_sheet(title="Ret")
    my_opx_append_rows(ws_ret, opx_df_to_rows_patched(portf_ret_df), styles=[date_style])

    output_std_rpt(wb, rpt.miss_ret_flg_rpt, 'MissRet')
    output_std_rpt(wb, rpt.wgt_rpt, 'Wgt')

    write_run_summary_xl(ws_sum, run_sum_df)
    write_fld_sum(ws_sum, style_sum_df, 'investment styles', 7)
    write_fld_sum(ws_sum, ccy_sum_df, 'currencies', starting_row=len(style_sum_df)+10)
    add_ret_chart(ws_sum,  ws_ret, "D7", portf_ret_df)

    if cfg.rpt.style_by_period:
        ws_st_cnt = wb.create_sheet(title="StyleC")
        my_opx_append_rows(ws_st_cnt, dataframe_to_rows(inv_stat_qtr_sum.cnt_df))

        fmt_arr = [None]*len(inv_stat_qtr_sum.frac_df.columns.levels[1])
        fmt_arr[1] = "0.0%"
        fmt_arr = fmt_arr * len(inv_stat_qtr_sum.frac_df.columns.levels[0])
        ws_st_frac = wb.create_sheet(title="StyleF")
        my_opx_append_rows(ws_st_frac, dataframe_to_rows(inv_stat_qtr_sum.frac_df)
            , hdr_row_cnt=len(inv_stat_qtr_sum.frac_df.columns.levels), hdr_col_cnt=1
            , number_formats=fmt_arr)


    wb.save("{}/{}".format(output_path, 'EH.xlsx'))
